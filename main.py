import openpyxl
from tkinter.filedialog import askdirectory
from tkinter import *
from tkinter import scrolledtext
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox




import os
global row

class Node ():
    character = None
    is_complete = False
    children = []
    row_number  = None

    def __init__(self,char):
        self.children = []
        self.character = char
        self.is_complete = False
        self.row_number = None


class Trie():
    Start_letters = []

    def new_word (self,type,new_word,List,row_number):
        length = len(new_word)
        if type == 0:
            list_of_objects = list(map(Node,new_word))
            self.Start_letters.append(list_of_objects[0])
            for i in range (length-1):
                list_of_objects[i].children.append(list_of_objects[i + 1])
            list_of_objects[-1].is_complete= True
            list_of_objects[-1].row_number = row_number
        else:
            myNode = List[0]
            start = List[1]
            if (start == length):
                myNode.is_complete = True
                myNode.row_number = row_number
                return
            else:
                list_of_objects = list(map(Node, new_word[start::]))
                myNode.children.append(list_of_objects[0])
                for i in range(length-start-1):
                    list_of_objects[i].children.append(list_of_objects[i + 1])
                list_of_objects[-1].is_complete = True
                list_of_objects[-1].row_number = row_number

    def search (self,word):
        flag = False
        length_of_word = len(word)
        index = 0
        for start in self.Start_letters:
            if word[0] == start.character:
                flag = True
                break
            index += 1
        if flag == False:
            return [0]
        else: #start letter mawgod
            counter = 1
            inside_Node = None
            inside_index = None
            flag3=False
            Node1= self.Start_letters[index]
            for k in range(1,length_of_word+1) :
                newflag = False
                for i in range(len(Node1.children)):
                    if k<length_of_word and Node1.children[i].character == word[k]:
                        flag3=True
                        inside_index = i
                        inside_Node = Node1
                        Node1 = Node1.children[i]
                        newflag = True
                        break
                if newflag == False:
                    break
            if flag3==True :
                Last_similar = inside_Node.children[inside_index]
            else:
                Last_similar = self.Start_letters[index]
            if Last_similar.is_complete ==True and k==(length_of_word):
                return  [Last_similar.row_number]
            else:
                return [-1,Last_similar,k] #if k =len --> da a5r harf

def mainpage():
    global root

    root = Tk()

    root.title("Main")
    root.geometry("300x300")

    button_1 = Button(root, text="Update your Database",command=update )
    button_1.grid(row=2,column=0,columnspan= 3, padx=10,pady=30)
    button_2 = Button(root, text="Search your Database",command=search)
    button_2.grid(row= 2,column=6,columnspan= 3, padx=10,pady=30)







def update():
    global row
    global root
    #root.destroy()
    myroot=Tk()
    myroot.geometry("300x300")
    myroot.title("Updating")
    progress = ttk.Progressbar(myroot,orient='horizontal',length=250, mode='determinate')
    progress.grid(padx=20,pady=20)
    mylabel = Label(myroot,text="Loading ... This may take a while")
    mylabel.grid(padx=20,pady=20)

    folder = askdirectory()
    List_of_files = list(filter(lambda x: x.endswith('.txt'), os.listdir(folder)))
    progress['maximum'] = len(List_of_files)

    for i in range(len(List_of_files)):
        progress['value'] = i
        progress.update()
        try:
            file = open(folder +'/'+ List_of_files[i], "r", encoding='UTF-8')
        except:
            file = open(folder +"/"+ List_of_files[i] , "r",encoding='latin-1”')

        f = set(file.read().split())
        for word in f:
            mylistt = myTrie.search(word)
            status = mylistt[0]
            if status > 0:  # y3ni l klma mwgoda f3ln w ba5od l row number b3tha
                if sheet_obj.cell(row=status, column=4).value != 7:
                    sheet_obj.cell(row=status, column=(4 + sheet_obj.cell(row=status, column=3).value)).value += List_of_files[i].split('.')[0] + "/"
                    sheet_obj.cell(row=status, column=4).value += 1
                else:
                    sheet_obj.cell(row=status, column=3).value += 1
                    sheet_obj.cell(row=status, column=4).value = 1
                    sheet_obj.cell(row=status, column=(4 + sheet_obj.cell(row=status, column=3).value)).value = List_of_files[i].split('.')[0] + "/"
            elif status == 0:  # el klma msh mwgoda 5ales w bdefha ml awl fl tie
                sheet_obj.cell(row=row, column=2).value = word
                sheet_obj.cell(row=row, column=3).value = 1
                sheet_obj.cell(row=row, column=4).value = 1
                myTrie.new_word(0, word, [], row)
                sheet_obj.cell(row=row, column=5).value = List_of_files[i].split('.')[0] + "/"
                row += 1
            elif status == -1:
                sheet_obj.cell(row=row, column=2).value = word
                sheet_obj.cell(row=row, column=3).value = 1
                sheet_obj.cell(row=row, column=4).value = 1
                myTrie.new_word(1, word, [mylistt[1], mylistt[2]], row)
                sheet_obj.cell(row=row, column=5).value = List_of_files[i].split('.')[0] + "/"
                row += 1
    sheet_obj.cell(row=1, column=1).value = row
    wb_obj.save("Reverse.xlsx")
    tkinter.messagebox.showinfo(title="Update Successfull", message="The Database was updated successfully")
    myroot.destroy()


def search():
    global Req_word
    #root.destroy()
    root2 = Tk()
    root2.title("Search")
    root2.geometry("400x300")
    label_1 = Label(root2, text="Please enter your search query")
    label_1.grid(row= 2,column=2,columnspan= 3, padx=8,pady=30)
    Req_word = Entry(root2, width=20)
    Req_word.grid(row= 2,column=6,columnspan= 5, padx=10,pady=30)
    button_1 = Button(root2, text="Search",command=query)
    button_1.grid(row=3, column=3, columnspan=7, padx=10, pady=20)



def query ():
    global Req_word
    word = Req_word.get()
    mylist=[]
    if myTrie.search(word)[0] > 0:
        number = int(sheet_obj.cell(row=myTrie.search(word)[0], column=3).value)
        final_num = (int(sheet_obj.cell(row=myTrie.search(word)[0], column=3).value) - 1) * 7 + int(
            sheet_obj.cell(row=myTrie.search(word)[0], column=4).value)
        #print("Your Search Query Occured " + str(final_num) + " times")
        #print("in files number:")
        for i in range(number):
             mylist.append(sheet_obj.cell(row=myTrie.search(word)[0], column=5 + i).value)
        scroll(mylist,final_num,word)

def result(mylist,number):
    result = Tk()
    result.title("Search Result")
    result.geometry("400x"+str(int(number*1.05)))

    if len(mylist) == 1:
        mylabel = Label(result, text=mylist[0])
        mylabel.grid(column=0,padx=5,pady = 3)
    elif len(mylist) == 2:
        mylabel = Label(result, text=mylist[0])
        mylabel.grid(row=0,column=0,padx=5,pady = 3)
        mylabel = Label(result, text=mylist[1])
        mylabel.grid(row=0,column=1,padx=5,pady = 3)

    else:
        for i in range(len(mylist)-2):
            mylabel = Label(result,text=mylist[i])
            mylabel.grid(row = i ,column=0,padx=5,pady = 3)
            mylabel = Label(result,text=mylist[i+1])
            mylabel.grid(row = i ,column=4,padx=5,pady=3)


def scroll(mylist,number,word):
    global file_name
    newroot = Tk()
    newroot.title("Search Results")
    newroot.geometry("500x500")

    mylabel = Label(newroot, text="Enter A File name:")
    mylabel.grid(row=0, column=0,columnspan=5, padx=0, pady=5)
    file_name = Entry(newroot, width=10)
    file_name.grid(row=0, column=4,columnspan=5, padx=0, pady=5)
    button = Button(newroot, text="View File",command=lambda :[show_File(file_name,word)])
    button.grid(row=1, column=2,columnspan=5,padx=0, pady=5)

    text_area = scrolledtext.ScrolledText(newroot ,wrap=tk.WORD,width=50,height=20,font = ("Times New Roman",13))
    text_area.grid(row=15,column=0,columnspan=5,padx=0, pady=5)
    text_area.insert(tk.INSERT,"Your Search Query occured:\n{} Times\n".format(number))
    text_area.insert(tk.INSERT,"In Files:\n")
    for text in mylist:
         text_area.insert(tk.INSERT,text+"\n")

def show_File(file,word):
    global file_name
    showroot=Tk()
    name_of_file=file_name.get()
    showroot.title("content of "+name_of_file+".txt")
    text_area = scrolledtext.ScrolledText(showroot, wrap=tk.WORD, width=50, height=8, font=("Times New Roman", 13))
    text_area.grid(row=15,column=0,columnspan=5,padx=0, pady=5)

    try:
        file = open("C:\\Users\\peter\\PycharmProjects\\DS\\questions\\"+name_of_file+".txt", "r", encoding='UTF-8')
    except:
        file = open("C:\\Users\\peter\\PycharmProjects\\DS\\questions\\"+name_of_file+".txt", "r", encoding='latin-1”')

    f = file.read().split()
    for text in f:
        if text == word:
            text_area.insert(tk.INSERT,text+" ",("A"))
            text_area.tag_config("A", foreground="blue")
        else:
            text_area.insert(tk.INSERT,text+" ")



myTrie = Trie()
wb_obj = openpyxl.load_workbook(filename="Reverse.xlsx")
sheet_obj = wb_obj.active

row = sheet_obj.cell(row=1, column=1).value
column=0


for counter in range(1, row):
    myword = sheet_obj.cell(row=counter, column=2).value
    status_list = myTrie.search(myword)
    status = status_list[0]
    if status == -1:
        myTrie.new_word(1,myword, [status_list[1],status_list[2]], counter)
    elif status ==0:
        myTrie.new_word(0, myword, [], counter)




mainpage()
root.mainloop()
































